import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font


def create_directory(directory):
    """
    Create reports directory if does not exist
    """
    if not os.path.exists('reports'):
        os.makedirs('reports')
        os.makedirs(os.path.join('reports', directory))
    elif not os.path.exists(os.path.join('reports', directory)):
        os.makedirs(os.path.join('reports', directory))


def create_report(results, algorithm, time, nodes, mu, avg_metrics):
    """
    Create Excel report with following data
    """
    print("Creating report for results")
    
    avg_nmi, avg_snmi, avg_ari, avg_vi, avg_purity, avg_fmeasure = avg_metrics

    wb = Workbook()
    ws = wb.active

    ws['A1'].value = 'Algorithm: {0}'.format(algorithm)
    ws['A3'].value = 'Number of Nodes'
    ws['B3'].value = 'Number of Edges'
    ws['C3'].value = 'Degree'
    ws['D3'].value = 'Gamma'
    ws['E3'].value = 'Beta'
    ws['F3'].value = 'MU'
    ws['G3'].value = 'GT Communities'
    ws['H3'].value = 'Communities Found'
    ws['I3'].value = 'NMI'
    ws['J3'].value = 'SNMI'
    ws['K3'].value = 'ARI'
    ws['L3'].value = 'VI'
    ws['M3'].value = 'Purity'
    ws['N3'].value = 'F - Measure'

    for col in range(12):
        ws.cell(row=1, column=col+1).font = Font(bold=True, size=12)

    for ind, result in enumerate(results):
        ws['A'+str(ind+4)].value = result[0]
        ws['B'+str(ind+4)].value = result[1]
        ws['C'+str(ind+4)].value = result[2]
        ws['D'+str(ind+4)].value = result[3]
        ws['E'+str(ind+4)].value = result[4]
        ws['F'+str(ind+4)].value = result[5]
        ws['G'+str(ind+4)].value = result[6]
        ws['H'+str(ind+4)].value = result[7]
        ws['I'+str(ind+4)].value = result[8]
        ws['J'+str(ind+4)].value = result[9]
        ws['K'+str(ind+4)].value = result[10]
        ws['L'+str(ind+4)].value = result[11]
        ws['M'+str(ind+4)].value = result[12]
        ws['N'+str(ind+4)].value = result[13]

    ws['I'+str(ind+6)].value = avg_nmi
    ws['J'+str(ind+6)].value = avg_snmi
    ws['K'+str(ind+6)].value = avg_ari
    ws['L'+str(ind+6)].value = avg_vi
    ws['M'+str(ind+6)].value = avg_purity
    ws['N'+str(ind+6)].value = avg_fmeasure


    mins, seconds = divmod(time, 60)
    ws['A' + str(ind+7)].value = "Executed time (min): {0}.{1}".format(mins, seconds)

    create_directory(algorithm)
    file_name = "{algorithm}_{nodes}_{mu}.xlsx".format(algorithm=algorithm, nodes=nodes, mu=mu)
    wb.save(os.path.join('reports', algorithm, file_name))
    wb.close

    print("Report generated")